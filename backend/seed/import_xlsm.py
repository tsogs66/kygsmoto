"""Import the shop's existing Excel workbook into the POS database.

Reads the KYGS workbook — inventory, categories, suppliers, service rates, the
critical-stock margins, the delisted list and the monthly trading history — and
writes it into SQLite. Safe to re-run: rows are matched on their natural keys
(SKU, category name, supplier code, service name) and updated in place.

    python -m backend.seed.import_xlsm "KYGS APRIL 2025.xlsm"
"""
import argparse
import os
import re
import sys
from datetime import date, datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from backend.app import db  # noqa: E402

INVENTORY_COLS = {
    "category": 0, "sku": 1, "description": 2, "supplier": 3, "unit_cost": 4,
    "opening_stock": 5, "retail_price": 12, "sales_qty": 13, "ending_stock": 15,
}
DAILY_FIRST_COL = 18   # Column S: day 1 of the month.
DAILY_LAST_COL = 49    # Column AW: day 31.


def _num(value, default=0.0):
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _text(value):
    return str(value).strip() if value is not None else ""


def detect_period(wb) -> str:
    """Work out which month the inventory sheet covers.

    The SALES log carries real dates, so trust it over the title cell, which the
    shop does not always update when they roll the sheet over.
    """
    try:
        for row in wb["SALES"].iter_rows(min_row=3, max_row=200, max_col=1, values_only=True):
            if isinstance(row[0], datetime):
                return row[0].strftime("%Y-%m")
    except KeyError:
        pass

    title = _text(wb["INVENTORY"]["A2"].value)
    match = re.match(r"([A-Z]+)\s+(\d{4})", title.upper())
    if match:
        try:
            return datetime.strptime(f"{match.group(1)[:3]} {match.group(2)}",
                                     "%b %Y").strftime("%Y-%m")
        except ValueError:
            pass
    return date.today().strftime("%Y-%m")


def import_infosheet(wb, stats):
    """Categories (name + code prefix), suppliers and service rates."""
    ws = wb["INFOSHEET"]
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=10, values_only=True))

    for row in rows:  # Column F = category name, column G = SKU prefix.
        name, prefix = _text(row[5]), _text(row[6])
        if not name:
            continue
        db.execute(
            "INSERT INTO categories(name, prefix) VALUES(?,?) "
            "ON CONFLICT(name) DO UPDATE SET prefix = excluded.prefix",
            (name.upper(), prefix.upper()),
        )
        stats["categories"] += 1

    for row in rows:  # Column D = supplier code.
        code = _text(row[3])
        if not code:
            continue
        db.execute(
            "INSERT INTO suppliers(code, name) VALUES(?,?) ON CONFLICT(code) DO NOTHING",
            (code.upper(), code.upper()),
        )
        stats["suppliers"] += 1

    for index, row in enumerate(rows, start=1):  # Columns A/B = service and fee.
        name, fee = _text(row[0]), _num(row[1], None)
        if not name or fee is None:
            continue
        existing = db.query_one("SELECT id FROM services WHERE name = ?", (name,))
        if existing:
            db.execute("UPDATE services SET fee = ? WHERE id = ?", (fee, existing["id"]))
        else:
            db.execute(
                "INSERT INTO services(code, name, fee) VALUES(?,?,?)",
                (f"SVC{index:03d}", name, fee),
            )
        stats["services"] += 1


def import_margins(wb):
    """Reorder points live on the CRITICAL sheet, keyed by description."""
    margins = {}
    try:
        ws = wb["CRITICAL"]
    except KeyError:
        return margins
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=5, values_only=True):
        description, margin = _text(row[1]), _num(row[2], None)
        if description and margin is not None:
            margins[description.upper()] = margin
    return margins


def import_inventory(wb, period, margins, stats):
    """The main item master, plus each item's sales for the sheet's month."""
    ws = wb["INVENTORY"]
    categories = {r["name"].upper(): r["id"] for r in db.query("SELECT id, name FROM categories")}
    suppliers = {r["code"].upper(): r["id"] for r in db.query("SELECT id, code FROM suppliers")}
    default_rop = float(db.get_setting("low_stock_default", "1"))
    seen_skus = {}

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row,
                            max_col=DAILY_LAST_COL, values_only=True):
        sku = _text(row[INVENTORY_COLS["sku"]]).upper()
        description = _text(row[INVENTORY_COLS["description"]])
        if not sku or not description:
            continue

        category_name = _text(row[INVENTORY_COLS["category"]]).upper()
        supplier_code = _text(row[INVENTORY_COLS["supplier"]]).upper()

        if category_name and category_name not in categories:
            cur = db.execute(
                "INSERT INTO categories(name, prefix) VALUES(?,?) ON CONFLICT(name) DO NOTHING",
                (category_name, category_name[:3]),
            )
            categories[category_name] = cur.lastrowid or db.query_one(
                "SELECT id FROM categories WHERE name = ?", (category_name,)
            )["id"]
        if supplier_code and supplier_code not in suppliers:
            cur = db.execute(
                "INSERT INTO suppliers(code, name) VALUES(?,?) ON CONFLICT(code) DO NOTHING",
                (supplier_code, supplier_code),
            )
            suppliers[supplier_code] = cur.lastrowid or db.query_one(
                "SELECT id FROM suppliers WHERE code = ?", (supplier_code,)
            )["id"]

        unit_cost = _num(row[INVENTORY_COLS["unit_cost"]])
        retail = _num(row[INVENTORY_COLS["retail_price"]])
        stock = _num(row[INVENTORY_COLS["ending_stock"]])
        reorder_point = margins.get(description.upper(), default_rop)

        existing = db.query_one("SELECT id, stock_qty FROM items WHERE sku = ?", (sku,))
        if sku in seen_skus:
            # The workbook occasionally reuses a code for two different parts.
            # The later row wins, but the clash is reported so the shop can fix it.
            stats["sku_collisions"].append(
                {"sku": sku, "kept": description, "replaced": seen_skus[sku]}
            )
        seen_skus[sku] = description

        if existing:
            db.execute(
                """UPDATE items SET description = ?, category_id = ?, supplier_id = ?,
                          unit_cost = ?, retail_price = ?, stock_qty = ?, reorder_point = ?,
                          updated_at = datetime('now')
                    WHERE id = ?""",
                (description, categories.get(category_name), suppliers.get(supplier_code),
                 unit_cost, retail, stock, reorder_point, existing["id"]),
            )
            item_id = existing["id"]
            stats["items_updated"] += 1
        else:
            cur = db.execute(
                """INSERT INTO items(sku, description, category_id, supplier_id, unit_cost,
                                     retail_price, stock_qty, reorder_point)
                   VALUES(?,?,?,?,?,?,?,?)""",
                (sku, description, categories.get(category_name),
                 suppliers.get(supplier_code), unit_cost, retail, stock, reorder_point),
            )
            item_id = cur.lastrowid
            stats["items_created"] += 1
            if stock:
                db.execute(
                    """INSERT INTO stock_moves(item_id, qty_delta, balance_after, move_type,
                                               ref_type, unit_cost, note)
                       VALUES(?,?,?,'opening','import',?,?)""",
                    (item_id, stock, stock, unit_cost, f"Imported opening stock {period}"),
                )

        # Month's demand: the daily grid sums to the sheet's SALES QUANTITY column.
        daily = [_num(v) for v in row[DAILY_FIRST_COL:DAILY_LAST_COL]]
        sold = sum(daily) or _num(row[INVENTORY_COLS["sales_qty"]])
        if sold > 0:
            db.execute(
                """INSERT INTO demand_history(item_id, period, qty, revenue, source)
                   VALUES(?,?,?,?,'workbook')
                   ON CONFLICT(item_id, period, source)
                   DO UPDATE SET qty = excluded.qty, revenue = excluded.revenue""",
                (item_id, period, sold, sold * retail),
            )
            stats["demand_rows"] += 1


def import_delisted(wb, stats):
    """Discontinued lines are kept, flagged, so history and reprints still resolve."""
    try:
        ws = wb["DELISTED"]
    except KeyError:
        return
    categories = {r["name"].upper(): r["id"] for r in db.query("SELECT id, name FROM categories")}
    suppliers = {r["code"].upper(): r["id"] for r in db.query("SELECT id, code FROM suppliers")}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=7, values_only=True):
        sku = _text(row[1]).upper()
        description = _text(row[2])
        if not sku or not description:
            continue
        if db.query_one("SELECT id FROM items WHERE sku = ?", (sku,)):
            db.execute("UPDATE items SET delisted = 1, active = 0 WHERE sku = ?", (sku,))
        else:
            db.execute(
                """INSERT INTO items(sku, description, category_id, supplier_id, unit_cost,
                                     retail_price, stock_qty, reorder_point, active, delisted)
                   VALUES(?,?,?,?,?,?,0,?,0,1)""",
                (sku, description, categories.get(_text(row[0]).upper()),
                 suppliers.get(_text(row[3]).upper()), _num(row[4]), _num(row[5]), _num(row[6])),
            )
        stats["delisted"] += 1


def import_monthly_history(wb, stats):
    """22 months of shop-level trading figures, for the management trend chart."""
    try:
        ws = wb["MONTHLY SALES"]
    except KeyError:
        return
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8, values_only=True):
        if not isinstance(row[0], datetime):
            continue  # Skips the AVERAGE / TOTAL footer rows.
        db.execute(
            """INSERT INTO monthly_summary(period, sales, income, expenses, net_profit,
                                           cash, merchandise, equity)
               VALUES(?,?,?,?,?,?,?,?)
               ON CONFLICT(period) DO UPDATE SET
                   sales = excluded.sales, income = excluded.income,
                   expenses = excluded.expenses, net_profit = excluded.net_profit,
                   cash = excluded.cash, merchandise = excluded.merchandise,
                   equity = excluded.equity""",
            (row[0].strftime("%Y-%m"), _num(row[1]), _num(row[2]), _num(row[3]),
             _num(row[4]), _num(row[5]), _num(row[6]), _num(row[7])),
        )
        stats["months"] += 1


def run(path: str) -> dict:
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("openpyxl is required for importing: pip install openpyxl")

    if not os.path.exists(path):
        raise SystemExit(f"Workbook not found: {path}")

    db.init_db()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    period = detect_period(wb)

    stats = {"categories": 0, "suppliers": 0, "services": 0, "items_created": 0,
             "items_updated": 0, "delisted": 0, "demand_rows": 0, "months": 0,
             "period": period, "sku_collisions": []}

    with db.transaction():
        import_infosheet(wb, stats)
        import_inventory(wb, period, import_margins(wb), stats)
        import_delisted(wb, stats)
        import_monthly_history(wb, stats)

    wb.close()
    return stats


def main():
    parser = argparse.ArgumentParser(description="Import the KYGS Excel workbook")
    parser.add_argument("workbook", help="Path to the .xlsm file")
    args = parser.parse_args()

    stats = run(args.workbook)
    collisions = stats.pop("sku_collisions", [])
    print("\nImport complete")
    for key, value in stats.items():
        print(f"  {key:<16} {value}")

    if collisions:
        print(f"\n  WARNING: {len(collisions)} item code(s) used twice in the workbook.")
        print("  The last row won; please give one of each pair a new code:")
        for clash in collisions:
            print(f"    {clash['sku']}: kept '{clash['kept']}'")
            print(f"    {' ' * len(clash['sku'])}  dropped '{clash['replaced']}'")
    totals = db.query_one(
        "SELECT COUNT(*) AS skus, ROUND(SUM(stock_qty * unit_cost), 2) AS value "
        "FROM items WHERE active = 1 AND delisted = 0"
    )
    print(f"\n  active SKUs      {totals['skus']}")
    print(f"  stock at cost    {totals['value']}\n")


if __name__ == "__main__":
    main()
